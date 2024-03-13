# -*- coding: utf-8 -*-
"""
Created on Thu Jan  2 12:05:36 2020

@author: Davide Laghi

Copyright 2021, the JADE Development Team. All rights reserved.

This file is part of JADE.

JADE is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

JADE is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with JADE.  If not, see <http://www.gnu.org/licenses/>.
"""
from copy import copy

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from xlsxwriter.utility import xl_rowcol_to_cell


def single_excel_writer(self, outpath, lib, testname, tallies, stats=None):
    """
    Produces single library summary excel file using XLSXwriter

    Parameters
    ----------
    outpath : path or str
        path to the output in Tests/Postprocessing.
    lib : str
        Shorthand representation of data library (i.e 00c, 31c).
    tallies : Dataframe
        Summary of tally tallies
    errors: Dataframe
        Errors on tally tallies
    stats: Dataframe
        Results of statistical checks

    Returns
    -------
    None
    """
    writer = pd.ExcelWriter(outpath, engine="xlsxwriter")

    # for df in (tallies, errors):
    # df.set_index("Zaid", inplace=True)

    startrow = 8
    startcol = 1
    max_len = 0
    max_width = 0
    df_positions = []

    for _, results in tallies.items():
        tally_len, tally_width = results["Value"].shape
        df_positions.append([startrow, startcol])
        # print(pd.Series(results["title"]))
        # pd.Series(results["title"]).to_excel(writer, startrow=startrow, startcol=startcol+1, sheet_name="Values", index=False, header=False)
        results["Value"].to_excel(
            writer, startrow=startrow + 1, startcol=startcol, sheet_name="Values"
        )
        results["Error"].to_excel(
            writer, startrow=startrow + 1, startcol=startcol, sheet_name="Errors"
        )
        startrow = startrow + tally_len + 3
        max_len = max_len + tally_len + 3
        if tally_width > max_width:
            max_width = tally_width

    wb = writer.book
    tal_sheet = writer.sheets["Values"]
    err_sheet = writer.sheets["Errors"]

    if stats is not None:
        # stats.set_index("Zaid", inplace=True)
        stats.to_excel(
            writer, startrow=8, startcol=0, sheet_name="Statistical Checks", index=False
        )
        stat_sheet = writer.sheets["Statistical Checks"]
        stats_len, stats_width = stats.shape

    # Formatting styles
    plain_format = wb.add_format({"bg_color": "#FFFFFF"})
    oob_format = wb.add_format(
        {
            "align": "center",
            "valign": "center",
            "bg_color": "#D9D9D9",
            "text_wrap": True,
        }
    )
    tally_format = wb.add_format({"bg_color": "#D9D9D9"})
    merge_format = wb.add_format({"align": "center", "valign": "center", "border": 2})
    title_merge_format = wb.add_format(
        {
            "font_size": "28",
            "align": "center",
            "valign": "center",
            "bold": True,
            "border": 2,
        }
    )
    subtitle_merge_format = wb.add_format(
        {
            "font_size": "16",
            "align": "center",
            "valign": "center",
            "bold": True,
            "border": 2,
        }
    )

    legend_text_format = wb.add_format(
        {"align": "center", "bg_color": "white", "border": 1}
    )
    red_cell_format = wb.add_format({"bg_color": "red", "border": 3})
    orange_cell_format = wb.add_format({"bg_color": "#FFC000", "border": 3})
    yellow_cell_format = wb.add_format({"bg_color": "#FFFF00", "border": 3})
    green_cell_format = wb.add_format({"bg_color": "#92D050", "border": 3})
    value_allzero_format = wb.add_format({"bg_color": "#EDEDED", "border": 3})
    value_belowzero_format = wb.add_format({"bg_color": "#FFC7CE", "border": 3})
    value_abovezero_format = wb.add_format({"bg_color": "#C6EFCE", "border": 3})

    scientific_format = wb.add_format({"num_format": "0.00E+00"})
    percent_format = wb.add_format({"num_format": "0.00%"})

    # tallies

    # Title Format
    tal_sheet.merge_range("B1:C2", "LIBRARY", subtitle_merge_format)
    tal_sheet.merge_range("D1:D2", lib, subtitle_merge_format)
    tal_sheet.merge_range(
        "B3:L8", "{} RESULTS RECAP: TALLIES".format(testname), title_merge_format
    )
    for tal in range(len(df_positions)):
        tal_sheet.merge_range(
            df_positions[tal][0],
            df_positions[tal][1] + 1,
            df_positions[tal][0],
            df_positions[tal][1] + 4,
            list(tallies.values())[tal]["title"],
            subtitle_merge_format,
        )
    # tal_sheet.merge_range("B8:C8", "ZAID", subtitle_merge_format)
    # tal_sheet.merge_range("D8:L8", "TALLY", subtitle_merge_format)

    # Freeze title
    tal_sheet.freeze_panes(8, 2)

    # out of bounds
    tal_sheet.set_column(0, 0, 4, oob_format)
    tal_sheet.set_column(max_width + 1, max_width + 1000, 18, oob_format)
    for i in range(9):
        tal_sheet.set_row(i, None, oob_format)
    for i in range(8 + max_len, max_len + 50):
        tal_sheet.set_row(i, None, oob_format)

    # Column widths
    tal_sheet.set_column(1, max_width + 1, 20)

    tal_sheet.conditional_format(
        10,
        1,
        8 + max_len,
        max_width + 1,
        {"type": "blanks", "format": oob_format},
    )
    tal_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": ">=",
            "value": 0,
            "format": scientific_format,
        },
    )
    tal_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "<",
            "value": 0,
            "format": scientific_format,
        },
    )

    # ERRORS

    # Title
    err_sheet.merge_range("B1:C2", "LIBRARY", subtitle_merge_format)
    err_sheet.merge_range("D1:D2", lib, subtitle_merge_format)
    err_sheet.merge_range(
        "B3:L8", "{} RESULTS RECAP: ERRORS".format(testname), title_merge_format
    )
    for tal in range(len(df_positions)):
        err_sheet.merge_range(
            df_positions[tal][0],
            df_positions[tal][1] + 1,
            df_positions[tal][0],
            df_positions[tal][1] + 4,
            list(tallies.values())[tal]["title"],
            subtitle_merge_format,
        )
    # err_sheet.merge_range("B8:C8", "ZAID", subtitle_merge_format)
    # err_sheet.merge_range("D8:L8", "TALLY", subtitle_merge_format)

    # Freeze title
    err_sheet.freeze_panes(8, 2)

    # out of bounds
    err_sheet.set_column(0, 0, 4, oob_format)
    err_sheet.set_column(max_width + 1, max_width + 1000, 18, oob_format)
    for i in range(9):
        err_sheet.set_row(i, None, oob_format)
    for i in range(8 + max_len, max_len + 50):
        err_sheet.set_row(i, None, oob_format)

    # Column widths
    err_sheet.set_column(1, max_width + 1, 20)

    # Legend
    err_sheet.merge_range("N3:O3", "LEGEND", merge_format)
    err_sheet.merge_range("N2:O2", "According to MCNP manual", oob_format)
    err_sheet.write("N4", "", red_cell_format)
    err_sheet.write("O4", "> 50%", legend_text_format)
    err_sheet.write("N5", "", orange_cell_format)
    err_sheet.write("O5", "20% ≤ 50%", legend_text_format)
    err_sheet.write("N6", "", yellow_cell_format)
    err_sheet.write("O6", "10% ≤ 20%", legend_text_format)
    err_sheet.write("N7", "", green_cell_format)
    err_sheet.write("O7", "< 10%", legend_text_format)

    # Conditional Formatting
    err_sheet.conditional_format(
        10,
        1,
        8 + max_len,
        max_width + 1,
        {"type": "blanks", "format": oob_format},
    )
    err_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "greater than",
            "value": 0.5,
            "format": red_cell_format,
        },
    )
    err_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 0.2,
            "maximum": 0.5,
            "format": orange_cell_format,
        },
    )
    err_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 0.1,
            "maximum": 0.2,
            "format": yellow_cell_format,
        },
    )
    err_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "less than",
            "value": -0.5,
            "format": red_cell_format,
        },
    )
    err_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.5,
            "maximum": -0.2,
            "format": orange_cell_format,
        },
    )
    err_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.2,
            "maximum": -0.1,
            "format": yellow_cell_format,
        },
    )
    err_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.1,
            "maximum": 0.1,
            "format": green_cell_format,
        },
    )
    err_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": ">=",
            "value": 0,
            "format": percent_format,
        },
    )
    err_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "<",
            "value": 0,
            "format": percent_format,
        },
    )

    # STAT CHECKS
    if stats is not None:
        # Title
        stat_sheet.merge_range("A1:B2", "LIBRARY", subtitle_merge_format)
        stat_sheet.merge_range("C1:C2", lib, subtitle_merge_format)
        stat_sheet.merge_range(
            "A3:C8",
            "10 MCNP Statistical Checks",
            title_merge_format,
        )
        # stat_sheet.merge_range("B8:C8", "ZAID", subtitle_merge_format)
        # stat_sheet.merge_range("D8:L8", "TALLY", subtitle_merge_format)

        # Freeze title
        stat_sheet.freeze_panes(9, 0)

        # out of bounds
        stat_sheet.set_column(stats_width + 2, stats_width + 20, 18, oob_format)
        for i in range(9):
            stat_sheet.set_row(i, None, oob_format)
        for i in range(9 + stats_len, stats_len + 50):
            stat_sheet.set_row(i, None, oob_format)

        # Column widths for errors, set up to 15th col by default to ensure title format correct
        stat_sheet.set_column(1, 1, 50)
        stat_sheet.set_column(0, 0, 30)
        stat_sheet.set_column(2, 2, 20)

        stat_sheet.conditional_format(
            9,
            2,
            8 + stats_len,
            2,
            {
                "type": "text",
                "criteria": "containing",
                "value": "Passed",
                "format": value_abovezero_format,
            },
        )
        stat_sheet.conditional_format(
            9,
            2,
            8 + stats_len,
            2,
            {
                "type": "text",
                "criteria": "containing",
                "value": "All zeros",
                "format": value_allzero_format,
            },
        )
        stat_sheet.conditional_format(
            9,
            2,
            8 + stats_len,
            2,
            {
                "type": "text",
                "criteria": "containing",
                "value": "Missed",
                "format": value_belowzero_format,
            },
        )

    wb.close()


def comp_excel_writer(self, outpath, lib_to_comp, testname, comps, abs_diffs, std_devs):
    """
    Produces library comparison excel file for general computation output
    using XLSXwriter

    Parameters
    ----------
    outpath : path or str
        path to the output in Tests/Postprocessing.
    name : str
        Shorthand representation of data libraries (i.e 00c-31c).
    final : Dataframe
       Percentage difference of all tallies
    absdiff: Dataframe
       Absolute difference between reference and target libraries
    std_dev: Dataframe
       Difference between reference and target library in terms of
       standard deviations from the mean of the reference library
    summary: Dataframe
       Contains total number of percentage difference in the column
       within certain bounds

    Returns
    -------
    None
    """
    writer = pd.ExcelWriter(outpath, engine="xlsxwriter")

    title = testname + " RESULTS RECAP: Comparison"
    startrow = 8
    startcol = 1
    max_len = 0
    max_width = 0
    df_positions = []

    for i in range(len(comps.keys())):
        comp_len, comp_width = list(comps.values())[i]["Value"].shape
        df_positions.append([startrow, startcol])
        list(comps.values())[i]["Value"].to_excel(
            writer,
            startrow=startrow + 1,
            startcol=startcol,
            sheet_name="Comparisons (%)",
        )
        list(std_devs.values())[i]["Value"].to_excel(
            writer,
            startrow=startrow + 1,
            startcol=startcol,
            sheet_name="Comparisons (std. dev.)",
        )
        list(abs_diffs.values())[i]["Value"].to_excel(
            writer,
            startrow=startrow + 1,
            startcol=startcol,
            sheet_name="Comparisons (abs. diff.)",
        )

        startrow = startrow + comp_len + 3
        max_len = max_len + comp_len + 3
        if comp_width > max_width:
            max_width = comp_width

    wb = writer.book
    comp_sheet = writer.sheets["Comparisons (%)"]
    std_dev_sheet = writer.sheets["Comparisons (std. dev.)"]
    absdiff_sheet = writer.sheets["Comparisons (abs. diff.)"]

    # Formatting styles
    plain_format = wb.add_format({"bg_color": "#FFFFFF"})
    oob_format = wb.add_format(
        {
            "align": "center",
            "valign": "center",
            "bg_color": "#D9D9D9",
            "text_wrap": True,
        }
    )
    tally_format = wb.add_format({"bg_color": "#D9D9D9"})
    merge_format = wb.add_format({"align": "center", "valign": "center", "border": 2})
    title_merge_format = wb.add_format(
        {
            "font_size": "36",
            "align": "center",
            "valign": "center",
            "bold": True,
            "border": 2,
        }
    )
    subtitle_merge_format = wb.add_format(
        {
            "font_size": "16",
            "align": "center",
            "valign": "center",
            "bold": True,
            "border": 2,
        }
    )
    legend_text_format = wb.add_format({"align": "center", "bg_color": "white"})
    red_cell_format = wb.add_format({"bg_color": "red"})
    orange_cell_format = wb.add_format({"bg_color": "orange"})
    yellow_cell_format = wb.add_format({"bg_color": "yellow"})
    green_cell_format = wb.add_format({"bg_color": "#A6D86E"})
    value_allzero_format = wb.add_format({"bg_color": "#EDEDED"})
    value_belowzero_format = wb.add_format({"bg_color": "#FFC7CE"})
    value_abovezero_format = wb.add_format({"bg_color": "#C6EFCE"})
    not_avail_format = wb.add_format({"bg_color": "#B8B8B8"})
    target_ref_format = wb.add_format({"bg_color": "#8465C5"})
    identical_format = wb.add_format({"bg_color": "#7ABD7E"})

    # COMPARISON

    # Title Format
    comp_sheet.merge_range("B1:C2", "LIBRARY", subtitle_merge_format)
    comp_sheet.merge_range("D1:D2", lib_to_comp, subtitle_merge_format)
    comp_sheet.merge_range(
        "B3:L8", "{} RESULTS RECAP: COMPARISON (%)".format(testname), title_merge_format
    )
    for tal in range(len(df_positions)):
        comp_sheet.merge_range(
            df_positions[tal][0],
            df_positions[tal][1] + 1,
            df_positions[tal][0],
            df_positions[tal][1] + 4,
            list(comps.values())[tal]["title"],
            subtitle_merge_format,
        )
    # comp_sheet.merge_range("B8:C8", "ZAID", subtitle_merge_format)
    # comp_sheet.merge_range("D8:L8", "TALLY", subtitle_merge_format)

    # Freeze title
    comp_sheet.freeze_panes(8, 2)

    # out of bounds
    comp_sheet.set_column(0, 0, 4, oob_format)
    comp_sheet.set_column(max_width + 2, max_width + 20, 18, oob_format)
    for i in range(9):
        comp_sheet.set_row(i, None, oob_format)
    for i in range(8 + max_len, max_len + 50):
        comp_sheet.set_row(i, None, oob_format)

    # Column widths for tallies, set up to 15th col to ensure title format correct
    comp_sheet.set_column(1, 14, 20)
    comp_sheet.set_column(1, max_width + 2, 20)

    # Row Heights
    comp_sheet.set_row(7, 31)
    # comp_sheet.set_row(8, 73.25)

    # Legend
    comp_sheet.merge_range("N3:O3", "LEGEND", merge_format)
    comp_sheet.write("N4", "", red_cell_format)
    comp_sheet.write("O4", ">|20|%", legend_text_format)
    comp_sheet.write("N5", "", orange_cell_format)
    comp_sheet.write("O5", "|20|%≤|10|%", legend_text_format)
    comp_sheet.write("N6", "", yellow_cell_format)
    comp_sheet.write("O6", "|10|%≤|5|%", legend_text_format)
    comp_sheet.write("N7", "", green_cell_format)
    comp_sheet.write("O7", "<|5|%", legend_text_format)

    comp_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {"type": "blanks", "format": oob_format},
    )
    comp_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {"type": "blanks", "format": plain_format},
    )
    comp_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Not Available",
            "format": not_avail_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Target = 0",
            "format": target_ref_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Reference = 0",
            "format": target_ref_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Identical",
            "format": identical_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "greater than",
            "value": 0.2,
            "format": red_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 0.1,
            "maximum": 0.2,
            "format": orange_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 0.05,
            "maximum": 0.1,
            "format": yellow_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "less than",
            "value": -0.2,
            "format": red_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.5,
            "maximum": -0.1,
            "format": orange_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.1,
            "maximum": -0.05,
            "format": yellow_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.05,
            "maximum": 0.05,
            "format": green_cell_format,
        },
    )
    # ABSOLUTE DIFFERENCE
    # Title
    absdiff_sheet.merge_range("B1:C2", "LIBRARY", subtitle_merge_format)
    absdiff_sheet.merge_range("D1:D2", lib_to_comp, subtitle_merge_format)
    absdiff_sheet.merge_range(
        "B3:L8",
        "{} RESULTS RECAP: COMPARISON (Absolute Difference)".format(testname),
        title_merge_format,
    )
    for tal in range(len(df_positions)):
        absdiff_sheet.merge_range(
            df_positions[tal][0],
            df_positions[tal][1] + 1,
            df_positions[tal][0],
            df_positions[tal][1] + 4,
            list(comps.values())[tal]["title"],
            subtitle_merge_format,
        )
    # absdiff_sheet.merge_range("B8:C8", "ZAID", subtitle_merge_format)
    # absdiff_sheet.merge_range("D8:L8", "TALLY", subtitle_merge_format)

    # Freeze title
    absdiff_sheet.freeze_panes(8, 2)

    # out of bounds
    absdiff_sheet.set_column(0, 0, 4, oob_format)
    absdiff_sheet.set_column(max_width + 2, max_width + 20, 18, oob_format)
    for i in range(9):
        absdiff_sheet.set_row(i, None, oob_format)
    for i in range(8 + max_len, max_len + 50):
        absdiff_sheet.set_row(i, None, oob_format)

    # Column widths for errors, set up to 15th col by default to ensure title format correct
    absdiff_sheet.set_column(1, 14, 20)
    absdiff_sheet.set_column(1, max_width + 2, 20)

    # Row Heights
    absdiff_sheet.set_row(7, 31)
    # absdiff_sheet.set_row(8, 73.25)

    # Legend
    absdiff_sheet.merge_range("N3:O3", "LEGEND", merge_format)
    absdiff_sheet.merge_range("N8:O8", "According to MCNP manual", oob_format)
    absdiff_sheet.write("N4", "", red_cell_format)
    absdiff_sheet.write("O4", "> 50%", legend_text_format)
    absdiff_sheet.write("N5", "", orange_cell_format)
    absdiff_sheet.write("O5", "20% ≤ 50%", legend_text_format)
    absdiff_sheet.write("N6", "", yellow_cell_format)
    absdiff_sheet.write("O6", "10% ≤ 20%", legend_text_format)
    absdiff_sheet.write("N7", "", green_cell_format)
    absdiff_sheet.write("O7", "< 10%", legend_text_format)

    # Conditional Formatting
    absdiff_sheet.conditional_format(
        10,
        1,
        8 + max_len,
        max_width + 1,
        {"type": "blanks", "format": oob_format},
    )
    # STANDARD DEVIATIONS

    # Title Format
    std_dev_sheet.merge_range("B1:C2", "LIBRARY", subtitle_merge_format)
    std_dev_sheet.merge_range("D1:D2", lib_to_comp, subtitle_merge_format)
    std_dev_sheet.merge_range(
        "B3:L8",
        "{} RESULTS RECAP: COMPARISON (Standard deviations from reference library)".format(
            testname
        ),
        title_merge_format,
    )
    for tal in range(len(df_positions)):
        std_dev_sheet.merge_range(
            df_positions[tal][0],
            df_positions[tal][1] + 1,
            df_positions[tal][0],
            df_positions[tal][1] + 4,
            list(comps.values())[tal]["title"],
            subtitle_merge_format,
        )
    # std_dev_sheet.merge_range("B8:C8", "ZAID", subtitle_merge_format)
    # std_dev_sheet.merge_range("D8:L8", "TALLY", subtitle_merge_format)

    # Freeze title
    std_dev_sheet.freeze_panes(8, 2)

    # out of bounds
    std_dev_sheet.set_column(0, 0, 4, oob_format)
    std_dev_sheet.set_column(max_width + 2, max_width + 20, 18, oob_format)
    for i in range(9):
        std_dev_sheet.set_row(i, None, oob_format)
    for i in range(8 + max_len, max_len + 50):
        std_dev_sheet.set_row(i, None, oob_format)

    # Column widths for tallies, set up to 15th col to ensure title format correct
    std_dev_sheet.set_column(1, 14, 20)
    std_dev_sheet.set_column(1, max_width + 2, 20)

    # Row Heights
    std_dev_sheet.set_row(7, 31)
    # std_dev_sheet.set_row(8, 73.25)

    # Legend
    std_dev_sheet.merge_range("N3:O3", "LEGEND", merge_format)
    std_dev_sheet.write("N4", "", red_cell_format)
    std_dev_sheet.write("O4", ">|3|%", legend_text_format)
    std_dev_sheet.write("N5", "", orange_cell_format)
    std_dev_sheet.write("O5", "|2|%≤|3|%", legend_text_format)
    std_dev_sheet.write("N6", "", yellow_cell_format)
    std_dev_sheet.write("O6", "|1|%≤|2|%", legend_text_format)
    std_dev_sheet.write("N7", "", green_cell_format)
    std_dev_sheet.write("O7", "<|1|%", legend_text_format)

    std_dev_sheet.conditional_format(
        10,
        1,
        8 + max_len,
        max_width + 1,
        {"type": "blanks", "format": oob_format},
    )
    std_dev_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {"type": "blanks", "format": plain_format},
    )
    std_dev_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Not Available",
            "format": not_avail_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Target = 0",
            "format": target_ref_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Reference = 0",
            "format": target_ref_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Identical",
            "format": identical_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "greater than",
            "value": 3,
            "format": red_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 2,
            "maximum": 3,
            "format": orange_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 1,
            "maximum": 2,
            "format": yellow_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "less than",
            "value": -3,
            "format": red_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -3,
            "maximum": -2,
            "format": orange_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -2,
            "maximum": -1,
            "format": yellow_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        2,
        8 + max_len,
        max_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -1,
            "maximum": 1,
            "format": green_cell_format,
        },
    )

    wb.close()


def sphere_single_excel_writer(self, outpath, lib, values, errors, stats=None):
    """
    Produces single library summary excel file for the sphere using XLSXwriter

    Parameters
    ----------
    outpath : path or str
        path to the output in Tests/Postprocessing.
    lib : str
        Shorthand representation of data library (i.e 00c, 31c).
    values : Dataframe
        Summary of tally values
    errors: Dataframe
        Errors on tally values
    stats: Dataframe
        Results of statistical checks

    Returns
    -------
    None
    """
    writer = pd.ExcelWriter(outpath, engine="xlsxwriter")

    wb = writer.book

    # Formatting styles
    plain_format = wb.add_format(
        {"align": "center", "valign": "center", "bg_color": "#FFFFFF"}
    )
    oob_format = wb.add_format(
        {
            "align": "center",
            "valign": "center",
            "bg_color": "#D9D9D9",
            "text_wrap": True,
        }
    )
    merge_format = wb.add_format(
        {"align": "center", "valign": "center", "border": 2, "text_wrap": True}
    )
    title_merge_format = wb.add_format(
        {
            "font_size": "36",
            "align": "center",
            "valign": "center",
            "bold": True,
            "border": 2,
            "text_wrap": True,
        }
    )
    subtitle_merge_format = wb.add_format(
        {
            "font_size": "16",
            "align": "center",
            "valign": "top",
            "bold": True,
            "border": 2,
            "text_wrap": True,
        }
    )
    subsubtitle_merge_format = wb.add_format(
        {
            "font_size": "12",
            "align": "center",
            "valign": "top",
            "bold": True,
            "border": 2,
            "text_wrap": True,
        }
    )
    format_with_alignment = wb.add_format(
        {
            "align": "center",
            "valign": "vcenter",
        }
    )

    legend_text_format = wb.add_format(
        {"align": "center", "bg_color": "white", "border": 1}
    )
    red_cell_format = wb.add_format({"bg_color": "red", "border": 3})
    orange_cell_format = wb.add_format({"bg_color": "#FFC000", "border": 3})
    yellow_cell_format = wb.add_format({"bg_color": "#FFFF00", "border": 3})
    green_cell_format = wb.add_format({"bg_color": "#92D050", "border": 3})
    value_allzero_format = wb.add_format({"bg_color": "#EDEDED", "border": 3})
    value_belowzero_format = wb.add_format({"bg_color": "#FFC7CE", "border": 3})
    value_abovezero_format = wb.add_format({"bg_color": "#C6EFCE", "border": 3})

    scientific_format = wb.add_format({"num_format": "0.00E+00"})
    percent_format = wb.add_format({"num_format": "0.00%"})

    # Populate Sheets
    # To wrap text can not overwrite dataframe formatting. https://stackoverflow.com/questions/42562977/xlsxwriter-text-wrap-not-working
    values.to_excel(
        writer, startrow=9, startcol=1, sheet_name="Values", index=False, header=False
    )
    val_sheet = writer.sheets["Values"]
    for col_num, value in enumerate(values.columns.values):
        val_sheet.write(8, col_num + 1, value, subsubtitle_merge_format)

    errors.to_excel(
        writer, startrow=9, startcol=1, sheet_name="Errors", index=False, header=False
    )
    err_sheet = writer.sheets["Errors"]
    for col_num, value in enumerate(errors.columns.values):
        err_sheet.write(8, col_num + 1, value, subsubtitle_merge_format)

    # Get shapes to define formatting bounds
    values_len, values_width = values.shape
    errors_len, errors_width = errors.shape

    if stats is not None:
        stats.to_excel(
            writer,
            startrow=9,
            startcol=1,
            sheet_name="Statistical Checks",
            index=False,
            header=False,
        )
        stat_sheet = writer.sheets["Statistical Checks"]
        stats_len, stats_width = stats.shape
        for col_num, value in enumerate(stats.columns.values):
            stat_sheet.write(8, col_num + 1, value, subsubtitle_merge_format)

    # Title Format
    val_sheet.merge_range("B1:C2", "LIBRARY", subtitle_merge_format)
    val_sheet.merge_range("D1:D2", lib, subtitle_merge_format)
    val_sheet.merge_range(
        "B3:Q7", "SPHERE LEAKAGE TEST RESULTS RECAP: VALUES", title_merge_format
    )
    val_sheet.merge_range("B8:C8", "ZAID", subtitle_merge_format)
    val_sheet.merge_range("D8:Q8", "TALLY", subtitle_merge_format)

    # Freeze title
    val_sheet.freeze_panes(9, 0)

    # out of bounds
    val_sheet.set_column(0, 0, 4, oob_format)
    val_sheet.set_column(values_width + 2, 1000, 18, oob_format)
    for i in range(9):
        val_sheet.set_row(i, None, oob_format)
    for i in range(9 + values_len, 1000):
        val_sheet.set_row(i, None, oob_format)

    # Column widths
    val_sheet.set_column(1, values_width + 2, 20)

    # Row Heights
    val_sheet.set_row(7, 31)
    val_sheet.set_row(8, 80)

    # Legend
    val_sheet.merge_range("S3:T3", "LEGEND", merge_format)
    val_sheet.write("S4", "", red_cell_format)
    val_sheet.write("T4", ">|5|%", legend_text_format)
    val_sheet.write("S5", "", orange_cell_format)
    val_sheet.write("T5", "|1|%≤|5|%", legend_text_format)
    val_sheet.write("S6", "", yellow_cell_format)
    val_sheet.write("T6", "|0.5|%≤|1|%", legend_text_format)
    val_sheet.write("S7", "", green_cell_format)
    val_sheet.write("T7", "<|0.5|%", legend_text_format)

    # Conditional Formatting
    val_sheet.conditional_format(
        9,
        3,
        8 + values_len,
        values_width + 1,
        {"type": "blanks", "format": plain_format},
    )
    val_sheet.conditional_format(
        9,
        3,
        8 + values_len,
        values_width + 1,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Value = 0",
            "format": value_allzero_format,
        },
    )
    val_sheet.conditional_format(
        9,
        3,
        8 + values_len,
        values_width + 1,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Value < 0",
            "format": value_belowzero_format,
        },
    )
    val_sheet.conditional_format(
        9,
        3,
        8 + values_len,
        values_width + 1,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Value > 0",
            "format": value_abovezero_format,
        },
    )
    val_sheet.conditional_format(
        9,
        3,
        8 + values_len,
        values_width + 1,
        {
            "type": "cell",
            "criteria": "greater than",
            "value": 0.05,
            "format": red_cell_format,
        },
    )
    val_sheet.conditional_format(
        9,
        3,
        8 + values_len,
        values_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 0.01,
            "maximum": 0.05,
            "format": orange_cell_format,
        },
    )
    val_sheet.conditional_format(
        9,
        3,
        8 + values_len,
        values_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 0.005,
            "maximum": 0.01,
            "format": yellow_cell_format,
        },
    )
    val_sheet.conditional_format(
        9,
        3,
        8 + values_len,
        values_width + 1,
        {
            "type": "cell",
            "criteria": "less than",
            "value": -0.05,
            "format": red_cell_format,
        },
    )
    val_sheet.conditional_format(
        9,
        3,
        8 + values_len,
        values_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.05,
            "maximum": -0.01,
            "format": orange_cell_format,
        },
    )
    val_sheet.conditional_format(
        9,
        3,
        8 + values_len,
        values_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.01,
            "maximum": -0.005,
            "format": yellow_cell_format,
        },
    )
    val_sheet.conditional_format(
        9,
        3,
        8 + values_len,
        values_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.005,
            "maximum": 0.005,
            "format": green_cell_format,
        },
    )

    # ERRORS

    # Title
    err_sheet.merge_range("B1:C2", "LIBRARY", subtitle_merge_format)
    err_sheet.merge_range("D1:D2", lib, subtitle_merge_format)
    err_sheet.merge_range(
        "B3:N7", "SPHERE LEAKAGE TEST RESULTS RECAP: ERRORS", title_merge_format
    )
    err_sheet.merge_range("B8:C8", "ZAID", subtitle_merge_format)
    err_sheet.merge_range("D8:N8", "TALLY", subtitle_merge_format)

    # Freeze title
    err_sheet.freeze_panes(9, 0)

    # out of bounds
    err_sheet.set_column(0, 0, 4, oob_format)
    err_sheet.set_column(errors_width, 1000, 18, oob_format)

    for i in range(9):
        err_sheet.set_row(i, None, oob_format)
    for i in range(9 + errors_len, 1000):
        err_sheet.set_row(i, None, oob_format)

    # Column widths
    err_sheet.set_column(1, errors_width, 20)

    # Row Heights
    err_sheet.set_row(7, 31)
    err_sheet.set_row(8, 80)

    # Legend
    err_sheet.merge_range("P3:Q3", "LEGEND", merge_format)
    err_sheet.merge_range("P8:Q8", "According to MCNP manual", oob_format)
    err_sheet.write("P4", "", red_cell_format)
    err_sheet.write("Q4", "> 50%", legend_text_format)
    err_sheet.write("P5", "", orange_cell_format)
    err_sheet.write("Q5", "20% ≤ 50%", legend_text_format)
    err_sheet.write("P6", "", yellow_cell_format)
    err_sheet.write("Q6", "10% ≤ 20%", legend_text_format)
    err_sheet.write("P7", "", green_cell_format)
    err_sheet.write("Q7", "< 10%", legend_text_format)

    # Conditional Formatting
    err_sheet.conditional_format(
        9,
        3,
        8 + errors_len,
        errors_width,
        {"type": "blanks", "format": oob_format},
    )
    err_sheet.conditional_format(
        9,
        3,
        8 + errors_len,
        errors_width,
        {"type": "cell", "criteria": ">", "value": 0.5, "format": red_cell_format},
    )
    err_sheet.conditional_format(
        9,
        3,
        8 + errors_len,
        errors_width,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 0.2,
            "maximum": 0.5,
            "format": orange_cell_format,
        },
    )
    err_sheet.conditional_format(
        9,
        3,
        8 + errors_len,
        errors_width,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 0.1,
            "maximum": 0.2,
            "format": yellow_cell_format,
        },
    )
    err_sheet.conditional_format(
        9,
        3,
        8 + errors_len,
        errors_width,
        {"type": "cell", "criteria": "<", "value": -0.5, "format": red_cell_format},
    )
    err_sheet.conditional_format(
        9,
        3,
        8 + errors_len,
        errors_width,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.5,
            "maximum": -0.2,
            "format": orange_cell_format,
        },
    )
    err_sheet.conditional_format(
        9,
        3,
        8 + errors_len,
        errors_width,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.2,
            "maximum": -0.1,
            "format": yellow_cell_format,
        },
    )
    err_sheet.conditional_format(
        9,
        3,
        8 + errors_len,
        errors_width,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.1,
            "maximum": 0.1,
            "format": green_cell_format,
        },
    )
    err_sheet.conditional_format(
        9,
        3,
        8 + errors_len,
        errors_width,
        {
            "type": "cell",
            "criteria": ">=",
            "value": 0,
            "format": percent_format,
        },
    )
    err_sheet.conditional_format(
        9,
        3,
        8 + errors_len,
        errors_width,
        {
            "type": "cell",
            "criteria": "<",
            "value": 0,
            "format": percent_format,
        },
    )

    # STAT CHECKS
    if stats is not None:
        # Title
        stat_sheet.merge_range("B1:C2", "LIBRARY", subtitle_merge_format)
        stat_sheet.merge_range("D1:D2", lib, subtitle_merge_format)
        stat_sheet.merge_range(
            "B3:N7",
            "SPHERE LEAKAGE TEST RESULTS RECAP: STATISTICAL CHECKS",
            title_merge_format,
        )
        stat_sheet.merge_range("B8:C8", "ZAID", subtitle_merge_format)
        stat_sheet.merge_range("D8:N8", "TALLY", subtitle_merge_format)

        # Freeze title
        stat_sheet.freeze_panes(9, 0)

        # out of bounds
        stat_sheet.set_column(0, 0, 4, oob_format)
        stat_sheet.set_column(stats_width, 1000, 18, oob_format)

        for i in range(9):
            stat_sheet.set_row(i, None, oob_format)
        for i in range(9 + stats_len, 1000):
            stat_sheet.set_row(i, None, oob_format)

        # Column widths
        stat_sheet.set_column(1, stats_width, 20)

        # Row Heights
        stat_sheet.set_row(7, 31)
        stat_sheet.set_row(8, 80)

        # Formatting
        stat_sheet.conditional_format(
            9,
            3,
            8 + stats_len,
            stats_width,
            {"type": "blanks", "format": plain_format},
        )
        stat_sheet.conditional_format(
            8,
            1,
            8,
            stats_width,
            {"format": subsubtitle_merge_format},
        )
        stat_sheet.conditional_format(
            9,
            3,
            8 + stats_len,
            stats_width,
            {
                "type": "text",
                "criteria": "containing",
                "value": "Passed",
                "format": value_abovezero_format,
            },
        )
        stat_sheet.conditional_format(
            9,
            3,
            8 + stats_len,
            stats_width,
            {
                "type": "text",
                "criteria": "containing",
                "value": "All zeros",
                "format": value_allzero_format,
            },
        )
        stat_sheet.conditional_format(
            9,
            3,
            8 + stats_len,
            stats_width,
            {
                "type": "text",
                "criteria": "containing",
                "value": "Missed",
                "format": value_belowzero_format,
            },
        )
        stat_sheet.conditional_format(
            9,
            3,
            8 + stats_len,
            stats_width,
            {
                "format": format_with_alignment,  # Apply the format with alignment
            },
        )

    wb.close()


def sphere_comp_excel_writer(self, outpath, name, final, absdiff, std_dev, summary):
    """
    Produces library comparison excel file for Sphere leakage using XLSXwriter

    Parameters
    ----------
    outpath : path or str
        path to the output in Tests/Postprocessing.
    name : str
        Shorthand representation of data libraries (i.e 00c-31c).
    final : Dataframe
       Percentage difference of all tallies
    absdiff: Dataframe
       Absolute difference between reference and target libraries
    std_dev: Dataframe
       Difference between reference and target library in terms of
       standard deviations from the mean of the reference library
    summary: Dataframe
       Contains total number of percentage difference in the column
       within certain bounds

    Returns
    -------
    None
    """
    writer = pd.ExcelWriter(
        outpath,
        engine="xlsxwriter",
        engine_kwargs={"options": {"strings_to_numbers": True}},
    )

    comp_len, comp_width = final.shape
    absdiff_len, absdiff_width = absdiff.shape
    std_dev_len, std_dev_width = std_dev.shape
    summ_len, summ_width = summary.shape

    # Insert DataFrame
    final.to_excel(writer, startrow=9, startcol=1, sheet_name="Comparison (%)")
    summary.to_excel(
        writer, startrow=comp_len + 12, startcol=2, sheet_name="Comparison (%)"
    )
    std_dev.to_excel(
        writer, startrow=9, startcol=1, sheet_name="Comparison (std. dev.)"
    )
    absdiff.to_excel(writer, startrow=9, startcol=1, sheet_name="Comparison (abs diff)")

    wb = writer.book
    comp_sheet = writer.sheets["Comparison (%)"]
    std_dev_sheet = writer.sheets["Comparison (std. dev.)"]
    absdiff_sheet = writer.sheets["Comparison (abs diff)"]

    # Formatting styles
    plain_format = wb.add_format({"bg_color": "#FFFFFF"})
    oob_format = wb.add_format(
        {
            "align": "center",
            "valign": "center",
            "bg_color": "#D9D9D9",
            "text_wrap": True,
        }
    )
    merge_format = wb.add_format({"align": "center", "valign": "center", "border": 2})
    title_merge_format = wb.add_format(
        {
            "font_size": "36",
            "align": "center",
            "valign": "center",
            "bold": True,
            "border": 2,
        }
    )
    subtitle_merge_format = wb.add_format(
        {
            "font_size": "16",
            "align": "center",
            "valign": "center",
            "bold": True,
            "border": 2,
            "text_wrap": True,
        }
    )
    legend_text_format = wb.add_format(
        {"align": "center", "bg_color": "white", "border": 1}
    )
    red_cell_format = wb.add_format({"bg_color": "red", "border": 3})
    orange_cell_format = wb.add_format({"bg_color": "#FFC000", "border": 3})
    yellow_cell_format = wb.add_format({"bg_color": "#FFFF00", "border": 3})
    green_cell_format = wb.add_format({"bg_color": "#92D050", "border": 3})
    not_avail_format = wb.add_format({"bg_color": "#B8B8B8", "border": 3})
    target_ref_format = wb.add_format({"bg_color": "#8465C5", "border": 3})
    identical_format = wb.add_format({"bg_color": "#7ABD7E", "border": 3})

    scientific_format = wb.add_format({"num_format": "0.00E+00"})
    percent_format = wb.add_format({"num_format": "0.00%"})

    """VALUES"""
    # Merged Cells
    comp_sheet.merge_range("B1:C2", "LIBRARY", subtitle_merge_format)
    comp_sheet.merge_range("D1:E2", name, subtitle_merge_format)
    comp_sheet.merge_range(
        "B3:T7", "SPHERE LEAKAGE % COMPARISON RECAP", title_merge_format
    )
    comp_sheet.merge_range("B8:C8", "ZAID", subtitle_merge_format)
    comp_sheet.merge_range("D8:T8", "TALLIES", subtitle_merge_format)
    comp_sheet.merge_range(
        "F1:L2",
        "Target library Vs Reference library\n(Reference-Target)/Reference",
        subtitle_merge_format,
    )
    comp_sheet.merge_range(
        "D9:I9", "Neutron Flux (Coarse energy bins) Tally n.12", subtitle_merge_format
    )
    comp_sheet.merge_range(
        "J9:O9", "Gamma  Flux (Coarse energy bins) Tally n.22", subtitle_merge_format
    )
    comp_sheet.merge_range("P9:T9", " ", subtitle_merge_format)
    comp_sheet.merge_range(
        11 + comp_len,
        3,
        11 + comp_len,
        20,
        "GLOBAL QUICK RESULT: % of cells per range of comparison differences",
        subtitle_merge_format,
    )
    # Freeze title
    comp_sheet.freeze_panes(10, 0)

    # out of bounds
    comp_sheet.set_column(0, 0, 4, oob_format)
    comp_sheet.set_column(comp_width, 1000, 4, oob_format)
    for i in range(9):
        comp_sheet.set_row(i, None, oob_format)
    for i in range(9 + comp_len, 1000):
        comp_sheet.set_row(i, None, oob_format)

    # Column widths
    comp_sheet.set_column(1, 14, 18)
    comp_sheet.set_column(1, comp_width + 5, 18)

    # Row Heights
    comp_sheet.set_row(0, 25, oob_format)
    comp_sheet.set_row(1, 25, oob_format)
    comp_sheet.set_row(7, 31, oob_format)
    comp_sheet.set_row(8, 40, oob_format)
    comp_sheet.set_row(9, 40, oob_format)

    # Legend
    comp_sheet.merge_range("V3:W3", "LEGEND", merge_format)
    comp_sheet.write("V4", "", red_cell_format)
    comp_sheet.write("W4", ">|20|%", legend_text_format)
    comp_sheet.write("V5", "", orange_cell_format)
    comp_sheet.write("W5", "|20|%≤|10|%", legend_text_format)
    comp_sheet.write("V6", "", yellow_cell_format)
    comp_sheet.write("W6", "|10|%≤|5|%", legend_text_format)
    comp_sheet.write("V7", "", green_cell_format)
    comp_sheet.write("W7", "<|5|%", legend_text_format)
    comp_sheet.write("V8", "Not Available", not_avail_format)
    comp_sheet.write("W8", "Both libs returned 0", legend_text_format)

    # Conditional Formatting
    comp_sheet.conditional_format(
        10,
        3,
        9 + comp_len,
        comp_width + 2,
        {"type": "blanks", "format": plain_format},
    )
    comp_sheet.conditional_format(
        10,
        3,
        9 + comp_len,
        comp_width + 2,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Not Available",
            "format": not_avail_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        3,
        9 + comp_len,
        comp_width + 2,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Target = 0",
            "format": target_ref_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        3,
        9 + comp_len,
        comp_width + 2,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Reference = 0",
            "format": target_ref_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        3,
        9 + comp_len,
        comp_width + 2,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Identical",
            "format": identical_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        3,
        9 + comp_len,
        comp_width + 2,
        {
            "type": "cell",
            "criteria": "greater than",
            "value": 0.2,
            "format": red_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        3,
        9 + comp_len,
        comp_width + 2,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 0.1,
            "maximum": 0.2,
            "format": orange_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        3,
        9 + comp_len,
        comp_width + 2,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 0.05,
            "maximum": 0.1,
            "format": yellow_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        3,
        9 + comp_len,
        comp_width + 2,
        {
            "type": "cell",
            "criteria": "less than",
            "value": -0.2,
            "format": red_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        3,
        9 + comp_len,
        comp_width + 2,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.5,
            "maximum": -0.1,
            "format": orange_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        3,
        9 + comp_len,
        comp_width + 2,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.1,
            "maximum": -0.05,
            "format": yellow_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        3,
        9 + comp_len,
        comp_width + 2,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.05,
            "maximum": 0.05,
            "format": green_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        3,
        9 + comp_len,
        comp_width + 2,
        {
            "type": "cell",
            "criteria": ">=",
            "value": 0,
            "format": percent_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        3,
        9 + comp_len,
        comp_width + 2,
        {
            "type": "cell",
            "criteria": "<",
            "value": 0,
            "format": percent_format,
        },
    )
    # Summary background formatting workaround using conditional formatting
    comp_sheet.conditional_format(
        12 + comp_len,
        3,
        comp_len + summ_len + 22,
        summ_width + 3,
        {
            "type": "cell",
            "criteria": ">=",
            "value": 0,
            "format": percent_format,
        },
    )
    comp_sheet.conditional_format(
        12 + comp_len,
        3,
        comp_len + summ_len + 22,
        summ_width + 3,
        {
            "type": "cell",
            "criteria": "<",
            "value": 0,
            "format": percent_format,
        },
    )
    # graded colour scale
    comp_sheet.conditional_format(
        12 + comp_len,
        3,
        comp_len + summ_len + 22,
        summ_width + 3,
        {
            "type": "2_color_scale",
            "min_color": "#FCFCFF",
            "max_color": "#F8696B",
        },
    )

    # Summary totals
    for row in range(13 + comp_len, comp_len + summ_len + 13):
        start = xl_rowcol_to_cell(row, 3)
        stop = xl_rowcol_to_cell(row, summ_width + 2)
        comp_sheet.write_formula(
            row,
            summ_width + 3,
            "=SUM({start}:{stop})/{len}".format(start=start, stop=stop, len=summ_width),
        )

    """STANDARD DEVIATIONS FROM MEAN"""
    std_dev_sheet.merge_range("B1:C2", "LIBRARY", subtitle_merge_format)
    std_dev_sheet.merge_range("D1:E2", name, subtitle_merge_format)
    std_dev_sheet.merge_range(
        "B3:T7",
        "SPHERE LEAKAGE COMPARISON (Standard deviations from reference library)",
        title_merge_format,
    )
    std_dev_sheet.merge_range("B8:C8", "ZAID", subtitle_merge_format)
    std_dev_sheet.merge_range("D8:T8", "TALLIES", subtitle_merge_format)
    std_dev_sheet.merge_range(
        "F1:L2",
        "Standard deviations from mean of reference library",
        subtitle_merge_format,
    )
    std_dev_sheet.merge_range(
        "D9:I9", "Neutron Flux (Coarse energy bins) Tally n.12", subtitle_merge_format
    )
    std_dev_sheet.merge_range(
        "J9:O9", "Gamma  Flux (Coarse energy bins) Tally n.22", subtitle_merge_format
    )
    std_dev_sheet.merge_range("P9:T9", " ", subtitle_merge_format)

    # Freeze title
    std_dev_sheet.freeze_panes(10, 0)

    # out of bounds
    std_dev_sheet.set_column(0, 0, 4, oob_format)
    std_dev_sheet.set_column(std_dev_width + 3, 1000, 4, oob_format)
    for i in range(9):
        std_dev_sheet.set_row(i, None, oob_format)
    for i in range(9 + std_dev_len, 1000):
        std_dev_sheet.set_row(i, None, oob_format)

    # Column widths for values, set up to 15th col to ensure title format correct

    std_dev_sheet.set_column(1, 14, 18)
    std_dev_sheet.set_column(1, std_dev_width + 5, 18)

    # Row Heights
    std_dev_sheet.set_row(0, 25, oob_format)
    std_dev_sheet.set_row(1, 25, oob_format)
    std_dev_sheet.set_row(7, 31, oob_format)
    std_dev_sheet.set_row(8, 40, oob_format)
    std_dev_sheet.set_row(9, 40, oob_format)

    # Legend
    std_dev_sheet.merge_range("V3:W3", "LEGEND", merge_format)
    std_dev_sheet.write("V4", "", red_cell_format)
    std_dev_sheet.write("W4", "3 < #σ", legend_text_format)
    std_dev_sheet.write("V5", "", orange_cell_format)
    std_dev_sheet.write("W5", "2 ≤ #σ ≤ 3", legend_text_format)
    std_dev_sheet.write("V6", "", yellow_cell_format)
    std_dev_sheet.write("W6", "1 ≤ #σ < 2", legend_text_format)
    std_dev_sheet.write("V7", "", green_cell_format)
    std_dev_sheet.write("W7", "#σ < 1", legend_text_format)

    # Conditional Formatting
    std_dev_sheet.conditional_format(
        10,
        3,
        9 + std_dev_len,
        std_dev_width + 2,
        {"type": "blanks", "format": plain_format},
    )
    std_dev_sheet.conditional_format(
        10,
        3,
        9 + std_dev_len,
        std_dev_width + 2,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Not Available",
            "format": not_avail_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        3,
        9 + std_dev_len,
        std_dev_width + 2,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Target = 0",
            "format": target_ref_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        3,
        9 + std_dev_len,
        std_dev_width + 2,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Reference = 0",
            "format": target_ref_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        3,
        9 + std_dev_len,
        std_dev_width + 2,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Identical",
            "format": identical_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        3,
        9 + std_dev_len,
        std_dev_width + 2,
        {
            "type": "cell",
            "criteria": "greater than",
            "value": 3,
            "format": red_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        3,
        9 + std_dev_len,
        std_dev_width + 2,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 3,
            "maximum": 2,
            "format": orange_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        3,
        9 + std_dev_len,
        std_dev_width + 2,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 2,
            "maximum": 1,
            "format": yellow_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        3,
        9 + std_dev_len,
        std_dev_width + 2,
        {
            "type": "cell",
            "criteria": "less than",
            "value": -3,
            "format": red_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        3,
        9 + std_dev_len,
        std_dev_width + 1,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -3,
            "maximum": -2,
            "format": orange_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        3,
        9 + std_dev_len,
        std_dev_width + 2,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -2,
            "maximum": -1,
            "format": yellow_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        3,
        9 + std_dev_len,
        std_dev_width + 2,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -1,
            "maximum": 1,
            "format": green_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": ">=",
            "value": 0,
            "format": scientific_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": "<",
            "value": 0,
            "format": scientific_format,
        },
    )

    """ABS DIFF"""
    # Merged Cells
    absdiff_sheet.merge_range("B1:C2", "LIBRARY", subtitle_merge_format)
    absdiff_sheet.merge_range("D1:E2", name, subtitle_merge_format)
    absdiff_sheet.merge_range(
        "B3:T7", "SPHERE LEAKAGE ABSOLUTE COMPARISON RECAP", title_merge_format
    )
    absdiff_sheet.merge_range("B8:C8", "ZAID", subtitle_merge_format)
    absdiff_sheet.merge_range("D8:T8", "TALLIES", subtitle_merge_format)
    absdiff_sheet.merge_range(
        "F1:L2",
        "Target library Vs Reference library\n(Reference-Target)",
        subtitle_merge_format,
    )
    absdiff_sheet.merge_range(
        "D9:I9", "Neutron Flux (Coarse energy bins) Tally n.12", subtitle_merge_format
    )
    absdiff_sheet.merge_range(
        "J9:O9", "Gamma  Flux (Coarse energy bins) Tally n.22", subtitle_merge_format
    )
    absdiff_sheet.merge_range("P9:T9", " ", subtitle_merge_format)

    # Freeze title
    absdiff_sheet.freeze_panes(10, 0)

    # out of bounds
    absdiff_sheet.set_column(0, 0, 4, oob_format)
    absdiff_sheet.set_column(absdiff_width + 3, 1000, 4, oob_format)
    for i in range(9):
        absdiff_sheet.set_row(i, None, oob_format)
    for i in range(10 + absdiff_len, 1000):
        absdiff_sheet.set_row(i, None, oob_format)

    # Column widths for values, set up to 15th col to ensure title format correct
    absdiff_sheet.set_column(1, 14, 18)
    absdiff_sheet.set_column(1, absdiff_width + 2, 18)

    # Row Heights
    absdiff_sheet.set_row(0, 25, oob_format)
    absdiff_sheet.set_row(1, 25, oob_format)
    absdiff_sheet.set_row(7, 31, oob_format)
    absdiff_sheet.set_row(8, 40, oob_format)
    absdiff_sheet.set_row(9, 40, oob_format)

    # Conditional Formatting

    absdiff_sheet.conditional_format(
        10,
        3,
        9 + absdiff_len,
        absdiff_width + 2,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Identical",
            "format": identical_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": ">=",
            "value": 0,
            "format": scientific_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": "<",
            "value": 0,
            "format": scientific_format,
        },
    )

    wb.close()


def sphere_sddr_single_excel_writer(outpath, lib, results, errors, stat_checks):
    """
    Produces single library summary excel file for Sphere SDDR using XLSXwriter

    Parameters
    ----------
    outpath : path or str
        path to the output in Tests/Postprocessing.
    lib : str
        Shorthand representation of data library (i.e 00c, 31c).
    results : Dataframe
        Summary of tally results
    errors: Dataframe
        Errors on tally errors
    stats_checks: Dataframe
        Results of statistical checks

    Returns
    -------
    None
    """
    writer = pd.ExcelWriter(outpath, engine="xlsxwriter")

    startrow = 9
    startcol = 0

    # Add the results column headers
    col_headers = [
        "PARENT",
        "PARENT NAME",
        " MT",
        "0s",
        "2.7h",
        "24h",
        "11.6d",
        "30d",
        "10y",
        "0s",
        "2.7h",
        "24h",
        "11.6d",
        "30d",
        "10y",
        "0s",
        "2.7h",
        "24h",
        "11.6d",
        "30d",
        "10y",
        "1E-6 MeV",
        "0.1 MeV",
        "1 MeV",
        "10 MeV",
        "20 MeV",
    ]

    max_len, max_width = results.shape
    results.to_excel(writer, startrow=startrow, startcol=startcol, sheet_name="Values")
    errors.to_excel(writer, startrow=startrow, startcol=startcol, sheet_name="Errors")
    wb = writer.book
    tal_sheet = writer.sheets["Values"]
    err_sheet = writer.sheets["Errors"]

    if stat_checks is not None:
        # stats.set_index("Zaid", inplace=True)
        stat_checks.to_excel(
            writer,
            startrow=startrow - 1,
            startcol=startcol,
            sheet_name="Statistical Checks",
        )
        stat_sheet = writer.sheets["Statistical Checks"]
        stats_len, stats_width = stat_checks.shape

    # Formatting styles
    plain_format = wb.add_format({"bg_color": "#FFFFFF"})
    oob_format = wb.add_format(
        {
            "align": "center",
            "valign": "center",
            "bg_color": "#D9D9D9",
            "text_wrap": True,
        }
    )
    tally_format = wb.add_format({"bg_color": "#D9D9D9"})
    merge_format = wb.add_format({"align": "center", "valign": "center", "border": 2})
    title_merge_format = wb.add_format(
        {
            "font_size": "36",
            "align": "center",
            "valign": "center",
            "bold": True,
            "border": 2,
        }
    )
    subtitle_merge_format = wb.add_format(
        {
            "font_size": "16",
            "align": "center",
            "valign": "center",
            "bold": True,
            "border": 2,
            "text_wrap": True,
        }
    )
    subsubtitle_merge_format = wb.add_format(
        {
            "font_size": "12",
            "align": "center",
            "valign": "center",
            "bold": True,
            "border": 2,
            "text_wrap": True,
        }
    )

    legend_text_format = wb.add_format({"align": "center", "bg_color": "white"})
    red_cell_format = wb.add_format({"bg_color": "red"})
    orange_cell_format = wb.add_format({"bg_color": "orange"})
    yellow_cell_format = wb.add_format({"bg_color": "yellow"})
    green_cell_format = wb.add_format({"bg_color": "#A6D86E"})
    value_allzero_format = wb.add_format({"bg_color": "#EDEDED"})
    value_belowzero_format = wb.add_format({"bg_color": "#FFC7CE"})
    value_abovezero_format = wb.add_format({"bg_color": "#C6EFCE"})

    # Title Format
    tal_sheet.merge_range("B1:C2", "LIBRARY", subtitle_merge_format)
    tal_sheet.merge_range("D1:D2", lib, subtitle_merge_format)
    tal_sheet.merge_range(
        "B3:AA7", "SPHERE SDDR TEST RESULTS RECAP: Values", title_merge_format
    )
    tal_sheet.merge_range("E8:AA8", "TALLY", subtitle_merge_format)
    tal_sheet.merge_range("B8:D9", "ZAID", subtitle_merge_format)
    tal_sheet.merge_range(
        "E9:J9",
        "Photon Flux at the External surface [p/cm^2/#s]",
        subtitle_merge_format,
    )
    tal_sheet.merge_range("K9:P9", "Shut Down Dose Rate [Sv/h]", subtitle_merge_format)
    tal_sheet.merge_range("Q9:V9", "Photon Heating [MeV/#s]", subtitle_merge_format)
    tal_sheet.merge_range(
        "W9:AA9", "Neutron Flux at the External surface", subtitle_merge_format
    )
    # Freeze title
    tal_sheet.freeze_panes(10, 0)

    start_col = 1
    # Add the results column headers
    for i, value in enumerate(col_headers):
        col_index = start_col + i
        tal_sheet.write(9, col_index, value, subsubtitle_merge_format)

    # out of bounds
    tal_sheet.set_column(0, 0, 4, oob_format)
    tal_sheet.set_column(max_width + 1, max_width + 20, 18, oob_format)
    for i in range(9):
        tal_sheet.set_row(i, None, oob_format)
    for i in range(10 + max_len, max_len + 200):
        tal_sheet.set_row(i, None, oob_format)

    # Column widths for tallies, set up to 26th col to ensure title format correct
    tal_sheet.set_column(4, 26, 8)
    tal_sheet.set_column(2, 2, 12)
    # tal_sheet.set_column(1, max_width + 2, 8)

    # Row Heights
    tal_sheet.set_row(8, 50, oob_format)
    tal_sheet.set_row(9, 45, oob_format)
    tal_sheet.set_row(1, 30, oob_format)

    # Conditional formatting for tally results.
    # tal_sheet.conditional_format(
    #    10,
    #    1,
    #    8 + max_len,
    #    max_width + 1,
    #    {"type": "blanks", "format": oob_format},
    # )
    tal_sheet.conditional_format(
        startrow + 1,
        startcol + 4,
        startrow + max_len,
        max_width + startcol,
        {
            "type": "cell",
            "criteria": "equal to",
            "value": 0,
            "format": oob_format,
        },
    )

    # ERRORS
    # Title
    err_sheet.merge_range("B1:C2", "LIBRARY", subtitle_merge_format)
    err_sheet.merge_range("D1:D2", lib, subtitle_merge_format)
    err_sheet.merge_range(
        "B3:X7", "SPHERE SDDR TEST RESULTS RECAP: Errors", title_merge_format
    )
    err_sheet.merge_range("E8:AA8", "TALLY", subtitle_merge_format)
    err_sheet.merge_range("B8:D9", "ZAID", subtitle_merge_format)
    err_sheet.merge_range(
        "E9:J9",
        "Photon Flux at the External surface [p/cm^2/#s]",
        subtitle_merge_format,
    )
    err_sheet.merge_range("K9:P9", "Shut Down Dose Rate [Sv/h]", subtitle_merge_format)
    err_sheet.merge_range("Q9:V9", "Photon Heating [MeV/#s]", subtitle_merge_format)
    err_sheet.merge_range(
        "W9:AA9", "Neutron Flux at the External surface", subtitle_merge_format
    )
    # Freeze title
    err_sheet.freeze_panes(10, 0)

    # Add the results column headers
    for i, value in enumerate(col_headers):
        col_index = start_col + i
        err_sheet.write(9, col_index, value, subsubtitle_merge_format)

    # out of bounds
    err_sheet.set_column(0, 0, 4, oob_format)
    err_sheet.set_column(max_width, max_width + 50, 18, oob_format)
    for i in range(9):
        err_sheet.set_row(i, None, oob_format)
    for i in range(8 + max_len, max_len + 50):
        err_sheet.set_row(i, None, oob_format)

    # Column widths for errors, set up to 15th col by default to ensure title format correct
    err_sheet.set_column(4, 26, 8)
    # err_sheet.set_column(1, max_width + 2, 8)

    # Row Heights
    err_sheet.set_row(8, 55)
    err_sheet.set_row(9, 45)
    err_sheet.set_row(1, 30, oob_format)
    # err_sheet.set_row(8, 73.25)

    # Legend
    err_sheet.merge_range("Y3:AA3", "LEGEND", merge_format)
    err_sheet.merge_range("Y2:AA2", "According to MCNP manual", oob_format)
    err_sheet.write("Y4", "", red_cell_format)
    err_sheet.merge_range("Z4:AA4", "> 50%", legend_text_format)
    err_sheet.write("Y5", "", orange_cell_format)
    err_sheet.merge_range("Z5:AA5", "20% ≤ 50%", legend_text_format)
    err_sheet.write("Y6", "", yellow_cell_format)
    err_sheet.merge_range("Z6:AA6", "10% ≤ 20%", legend_text_format)
    err_sheet.write("Y7", "", green_cell_format)
    err_sheet.merge_range("Z7:AA7", "< 10%", legend_text_format)

    # Conditional Formatting
    err_sheet.conditional_format(
        startrow + 1,
        startcol + 4,
        startrow + max_len,
        max_width + startcol,
        {"type": "blanks", "format": oob_format},
    )
    err_sheet.conditional_format(
        startrow + 1,
        startcol + 4,
        startrow + max_len,
        max_width + startcol,
        {
            "type": "cell",
            "criteria": "greater than",
            "value": 0.5,
            "format": red_cell_format,
        },
    )
    err_sheet.conditional_format(
        startrow + 1,
        startcol + 4,
        startrow + max_len,
        max_width + startcol,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 0.2,
            "maximum": 0.5,
            "format": orange_cell_format,
        },
    )
    err_sheet.conditional_format(
        startrow + 1,
        startcol + 4,
        startrow + max_len,
        max_width + startcol,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 0.1,
            "maximum": 0.2,
            "format": yellow_cell_format,
        },
    )
    err_sheet.conditional_format(
        startrow + 1,
        startcol + 4,
        startrow + max_len,
        max_width + startcol,
        {
            "type": "cell",
            "criteria": "less than",
            "value": -0.5,
            "format": red_cell_format,
        },
    )
    err_sheet.conditional_format(
        startrow + 1,
        startcol + 4,
        startrow + max_len,
        max_width + startcol,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.5,
            "maximum": -0.2,
            "format": orange_cell_format,
        },
    )
    err_sheet.conditional_format(
        startrow + 1,
        startcol + 4,
        startrow + max_len,
        max_width + startcol,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.2,
            "maximum": -0.1,
            "format": yellow_cell_format,
        },
    )
    err_sheet.conditional_format(
        startrow + 1,
        startcol + 4,
        startrow + max_len,
        max_width + startcol,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.1,
            "maximum": 0.1,
            "format": green_cell_format,
        },
    )

    # STAT CHECKS
    if stat_checks is not None:
        # Title
        stat_sheet.merge_range("B1:C2", "LIBRARY", subtitle_merge_format)
        stat_sheet.merge_range("D1:D2", lib, subtitle_merge_format)
        stat_sheet.merge_range(
            "B3:P7",
            "SPHERE SDDR TEST RESULTS RECAP: 10 MCNP statistical checks",
            title_merge_format,
        )
        stat_sheet.merge_range("B8:D8", "ZAID", subtitle_merge_format)
        stat_sheet.merge_range("E8:P8", "TALLY", subtitle_merge_format)

        # Freeze title
        stat_sheet.freeze_panes(9, 0)

        # out of bounds
        stat_sheet.set_column(0, 0, 4, oob_format)
        stat_sheet.set_column(stats_width + 1, stats_width + 50, 18, oob_format)
        for i in range(9):
            stat_sheet.set_row(i, None, oob_format)
        for i in range(9 + stats_len, stats_len + 50):
            stat_sheet.set_row(i, None, oob_format)

        stat_sheet.set_column(1, 3, 8)
        stat_sheet.set_column(1, stats_width, 20)

        # Row Heights
        stat_sheet.set_row(7, 31)
        stat_sheet.set_row(8, 73.25)

        # Formatting
        stat_sheet.conditional_format(
            startrow,
            startcol,
            startrow + stats_len - 1,
            stats_width + startcol,
            {"type": "blanks", "format": plain_format},
        )
        stat_sheet.conditional_format(
            startrow - 1,
            startcol + 1,
            startrow - 1,
            stats_width + startcol,
            {
                "type": "text",
                "criteria": "containsText",
                format: subsubtitle_merge_format,
            },
        )
        stat_sheet.conditional_format(
            startrow - 1,
            startcol,
            startrow - 1 + stats_len,
            stats_width + startcol,
            {
                "type": "text",
                "criteria": "containing",
                "value": "Passed",
                "format": value_abovezero_format,
            },
        )
        stat_sheet.conditional_format(
            startrow - 1,
            startcol,
            startrow - 1 + stats_len,
            stats_width + startcol,
            {
                "type": "text",
                "criteria": "containing",
                "value": "All zeros",
                "format": value_allzero_format,
            },
        )
        stat_sheet.conditional_format(
            startrow - 1,
            startcol,
            startrow - 1 + stats_len,
            stats_width + startcol,
            {
                "type": "text",
                "criteria": "containing",
                "value": "Missed",
                "format": value_belowzero_format,
            },
        )

    wb.close()


def sphere_sddr_comp_excel_writer(
    outpath, name, final, absdiff, std_dev, single_pp_files
):
    """
    Produces library comparison excel file for Sphere SDDR using XLSXwriter

    Parameters
    ----------
    outpath : path or str
        path to the output in Tests/Postprocessing.
    name : str
        Shorthand representation of data libraries (i.e 00c-31c).
    final : Dataframe
       Percentage difference of all tallies
    absdiff: Dataframe
       Absolute difference between reference and target libraries
    std_dev: Dataframe
       Difference between reference and target library in terms of
       standard deviations from the mean of the reference library


    Returns
    -------
    None
    """
    writer = pd.ExcelWriter(
        outpath,
        engine="xlsxwriter",
        engine_kwargs={"options": {"strings_to_numbers": True}},
    )
    comp_len, comp_width = final.shape
    absdiff_len, absdiff_width = absdiff.shape
    std_dev_len, std_dev_width = std_dev.shape

    # Insert DataFrame
    final.to_excel(writer, startrow=9, startcol=1, sheet_name="Comparison (%)")
    std_dev.to_excel(
        writer, startrow=9, startcol=1, sheet_name="Comparison (std. dev.)"
    )
    absdiff.to_excel(
        writer, startrow=9, startcol=1, sheet_name="Comparison (Abs diff.)"
    )
    wb = writer.book
    comp_sheet = writer.sheets["Comparison (%)"]
    std_dev_sheet = writer.sheets["Comparison (std. dev.)"]
    absdiff_sheet = writer.sheets["Comparison (Abs diff.)"]

    # Add the results column headers
    col_headers = [
        "PARENT",
        "PARENT NAME",
        " MT",
        "0s",
        "2.7h",
        "24h",
        "11.6d",
        "30d",
        "10y",
        "0s",
        "2.7h",
        "24h",
        "11.6d",
        "30d",
        "10y",
        "0s",
        "2.7h",
        "24h",
        "11.6d",
        "30d",
        "10y",
        "1E-6 MeV",
        "0.1 MeV",
        "1 MeV",
        "10 MeV",
        "20 MeV",
    ]

    # Formatting styles
    plain_format = wb.add_format({"bg_color": "#FFFFFF"})
    oob_format = wb.add_format(
        {
            "align": "center",
            "valign": "center",
            "bg_color": "#D9D9D9",
            "text_wrap": True,
        }
    )
    tally_format = wb.add_format({"bg_color": "#D9D9D9"})
    merge_format = wb.add_format({"align": "center", "valign": "center", "border": 2})
    title_merge_format = wb.add_format(
        {
            "font_size": "36",
            "align": "center",
            "valign": "center",
            "bold": True,
            "border": 2,
        }
    )
    subtitle_merge_format = wb.add_format(
        {
            "font_size": "16",
            "align": "center",
            "valign": "center",
            "bold": True,
            "border": 2,
            "text_wrap": True,
        }
    )
    subsubtitle_merge_format = wb.add_format(
        {
            "font_size": "12",
            "align": "center",
            "valign": "center",
            "bold": True,
            "border": 2,
            "text_wrap": True,
        }
    )
    legend_text_format = wb.add_format({"align": "center", "bg_color": "white"})
    red_cell_format = wb.add_format({"align": "center", "bg_color": "FF6961"})
    orange_cell_format = wb.add_format({"align": "center", "bg_color": "FFB54C"})
    yellow_cell_format = wb.add_format({"align": "center", "bg_color": "F8D66D"})
    green_cell_format = wb.add_format({"align": "center", "bg_color": "#8CD47E"})
    not_avail_format = wb.add_format({"align": "center", "bg_color": "#B8B8B8"})
    target_ref_format = wb.add_format({"align": "center", "bg_color": "#8465C5"})
    identical_format = wb.add_format({"align": "center", "bg_color": "#7ABD7E"})

    scientific_format = wb.add_format({"num_format": "0.00E+00"})
    percent_format = wb.add_format({"num_format": "0.00%"})

    # Title Format
    comp_sheet.merge_range(
        "H1:N2",
        "Target library vs Reference library, (Reference - Target)/Target",
        subtitle_merge_format,
    )
    comp_sheet.merge_range("B1:D2", "LIBRARY", subtitle_merge_format)
    comp_sheet.merge_range("E1:G2", name, subtitle_merge_format)
    comp_sheet.merge_range(
        "B3:X7", "SPHERE SDDR % COMPARISON RECAP", title_merge_format
    )
    comp_sheet.merge_range("E8:AA8", "TALLY", subtitle_merge_format)
    comp_sheet.merge_range("B8:D9", "ZAID", subtitle_merge_format)
    comp_sheet.merge_range(
        "E9:J9",
        "Photon Flux at the External surface",
        subtitle_merge_format,
    )
    comp_sheet.merge_range("K9:P9", "Shut Down Dose Rate", subtitle_merge_format)
    comp_sheet.merge_range("Q9:V9", "Photon Heating", subtitle_merge_format)
    comp_sheet.merge_range(
        "W9:AA9", "Neutron Flux at the External surface", subtitle_merge_format
    )

    # Freeze title
    comp_sheet.freeze_panes(9, 0)

    start_col = 1
    # Add the results column headers
    for i, value in enumerate(col_headers):
        col_index = start_col + i
        comp_sheet.write(9, col_index, value, subsubtitle_merge_format)

    # out of bounds
    comp_sheet.set_column(0, 0, 4, oob_format)
    comp_sheet.set_column(comp_width + 3, 1000, 4, oob_format)
    for i in range(9):
        comp_sheet.set_row(i, None, oob_format)
    for i in range(10 + comp_len, 1000):
        comp_sheet.set_row(i, None, oob_format)

    # Column widths for values
    comp_sheet.set_column(2, 2, 15)
    comp_sheet.set_column(1, comp_width + 3, 12)

    # Row Heights
    comp_sheet.set_row(0, 25, oob_format)
    comp_sheet.set_row(1, 25, oob_format)
    comp_sheet.set_row(7, 31, oob_format)
    comp_sheet.set_row(8, 40, oob_format)
    comp_sheet.set_row(9, 40, oob_format)

    # Legend
    comp_sheet.merge_range("Y3:AA3", "LEGEND", merge_format)
    comp_sheet.merge_range("Y2:AA2", "According to MCNP manual", oob_format)
    comp_sheet.write("Y4", "", red_cell_format)
    comp_sheet.merge_range("Z4:AA4", "> 50%", legend_text_format)
    comp_sheet.write("Y5", "", orange_cell_format)
    comp_sheet.merge_range("Z5:AA5", "20% ≤ 50%", legend_text_format)
    comp_sheet.write("Y6", "", yellow_cell_format)
    comp_sheet.merge_range("Z6:AA6", "10% ≤ 20%", legend_text_format)
    comp_sheet.write("Y7", "", green_cell_format)
    comp_sheet.merge_range("Z7:AA7", "< 10%", legend_text_format)

    # Conditional Formatting
    comp_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {"type": "blanks", "format": plain_format},
    )
    comp_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Not Available",
            "format": not_avail_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": ">=",
            "value": 0,
            "format": percent_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": "<",
            "value": 0,
            "format": percent_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Target = 0",
            "format": target_ref_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Reference = 0",
            "format": target_ref_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Identical",
            "format": identical_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": "greater than",
            "value": 0.2,
            "format": red_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 0.1,
            "maximum": 0.2,
            "format": orange_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 0.05,
            "maximum": 0.1,
            "format": yellow_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": "less than",
            "value": -0.2,
            "format": red_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.5,
            "maximum": -0.1,
            "format": orange_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.1,
            "maximum": -0.05,
            "format": yellow_cell_format,
        },
    )
    comp_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -0.05,
            "maximum": 0.05,
            "format": green_cell_format,
        },
    )

    """STANDARD DEVIATIONS FROM MEAN"""
    # Title Format
    std_dev_sheet.merge_range(
        "H1:N2",
        "Standard deviations from mean of reference lib",
        subtitle_merge_format,
    )
    std_dev_sheet.merge_range("B1:D2", "LIBRARY", subtitle_merge_format)
    std_dev_sheet.merge_range("E1:G2", name, subtitle_merge_format)
    std_dev_sheet.merge_range(
        "B3:X7", "SPHERE SDDR TEST COMPARISON: (std. dev.)", title_merge_format
    )
    std_dev_sheet.merge_range("E8:AA8", "TALLY", subtitle_merge_format)
    std_dev_sheet.merge_range("B8:D9", "ZAID", subtitle_merge_format)
    std_dev_sheet.merge_range(
        "E9:J9",
        "Photon Flux at the External surface",
        subtitle_merge_format,
    )
    std_dev_sheet.merge_range("K9:P9", "Shut Down Dose Rate", subtitle_merge_format)
    std_dev_sheet.merge_range("Q9:V9", "Photon Heating", subtitle_merge_format)
    std_dev_sheet.merge_range(
        "W9:AA9", "Neutron Flux at the External surface", subtitle_merge_format
    )
    # Freeze title
    std_dev_sheet.freeze_panes(9, 0)

    start_col = 1
    # Add the results column headers
    for i, value in enumerate(col_headers):
        col_index = start_col + i
        std_dev_sheet.write(9, col_index, value, subsubtitle_merge_format)

    # out of bounds
    std_dev_sheet.set_column(0, 0, 4, oob_format)
    std_dev_sheet.set_column(std_dev_width + 3, 1000, 4, oob_format)
    for i in range(9):
        std_dev_sheet.set_row(i, None, oob_format)
    for i in range(10 + std_dev_len, 1000):
        std_dev_sheet.set_row(i, None, oob_format)

    # Column widths for values
    std_dev_sheet.set_column(2, 2, 15)
    std_dev_sheet.set_column(1, std_dev_width + 3, 12)

    # Row Heights
    std_dev_sheet.set_row(0, 25, oob_format)
    std_dev_sheet.set_row(1, 25, oob_format)
    std_dev_sheet.set_row(7, 31, oob_format)
    std_dev_sheet.set_row(8, 40, oob_format)
    std_dev_sheet.set_row(9, 40, oob_format)

    # Legend
    std_dev_sheet.merge_range("Y3:AA3", "LEGEND", merge_format)
    std_dev_sheet.merge_range("Y2:AA2", "", oob_format)
    std_dev_sheet.write("Y4", "", red_cell_format)
    std_dev_sheet.merge_range("Z4:AA4", "> 3σ", legend_text_format)
    std_dev_sheet.write("Y5", "", orange_cell_format)
    std_dev_sheet.merge_range("Z5:AA5", "2σ ≤ 3σ", legend_text_format)
    std_dev_sheet.write("Y6", "", yellow_cell_format)
    std_dev_sheet.merge_range("Z6:AA6", "1σ ≤ 2σ", legend_text_format)
    std_dev_sheet.write("Y7", "", green_cell_format)
    std_dev_sheet.merge_range("Z7:AA7", "< 1σ", legend_text_format)

    # Conditional Formatting
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + std_dev_len,
        std_dev_width + 3,
        {"type": "blanks", "format": plain_format},
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": ">=",
            "value": 0,
            "format": scientific_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": "<",
            "value": 0,
            "format": scientific_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + std_dev_len,
        std_dev_width + 3,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Not Available",
            "format": not_avail_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + std_dev_len,
        std_dev_width + 3,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Target = 0",
            "format": target_ref_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + std_dev_len,
        std_dev_width + 3,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Reference = 0",
            "format": target_ref_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + std_dev_len,
        std_dev_width + 3,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Identical",
            "format": identical_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + std_dev_len,
        std_dev_width + 3,
        {
            "type": "cell",
            "criteria": "greater than",
            "value": 3,
            "format": red_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + std_dev_len,
        std_dev_width + 3,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 3,
            "maximum": 2,
            "format": orange_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + std_dev_len,
        std_dev_width + 3,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": 2,
            "maximum": 1,
            "format": yellow_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + std_dev_len,
        std_dev_width + 3,
        {
            "type": "cell",
            "criteria": "less than",
            "value": -3,
            "format": red_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + std_dev_len,
        std_dev_width + 3,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -3,
            "maximum": -2,
            "format": orange_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + std_dev_len,
        std_dev_width + 3,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -2,
            "maximum": -1,
            "format": yellow_cell_format,
        },
    )
    std_dev_sheet.conditional_format(
        10,
        4,
        9 + std_dev_len,
        std_dev_width + 3,
        {
            "type": "cell",
            "criteria": "between",
            "minimum": -1,
            "maximum": 1,
            "format": green_cell_format,
        },
    )

    """ABS DIFF"""
    # Title
    absdiff_sheet.merge_range(
        "H1:N2",
        "Target library vs Reference library, (Reference - Target)",
        subtitle_merge_format,
    )
    absdiff_sheet.merge_range("B1:D2", "LIBRARY", subtitle_merge_format)
    absdiff_sheet.merge_range("E1:G2", name, subtitle_merge_format)
    absdiff_sheet.merge_range(
        "B3:AA7",
        "SPHERE SDDR TEST COMPARISON: (ABSOLUTE DIFFERENCE)",
        title_merge_format,
    )
    absdiff_sheet.merge_range("E8:AA8", "TALLY", subtitle_merge_format)
    absdiff_sheet.merge_range("B8:D9", "ZAID", subtitle_merge_format)
    absdiff_sheet.merge_range(
        "E9:J9",
        "Photon Flux at the External surface [p/cm^2/#s]",
        subtitle_merge_format,
    )
    absdiff_sheet.merge_range(
        "K9:P9", "Shut Down Dose Rate [Sv/h]", subtitle_merge_format
    )
    absdiff_sheet.merge_range("Q9:V9", "Photon Heating [MeV/#s]", subtitle_merge_format)
    absdiff_sheet.merge_range(
        "W9:AA9", "Neutron Flux at the External surface", subtitle_merge_format
    )
    # Freeze title
    absdiff_sheet.freeze_panes(9, 0)

    start_col = 1
    # Add the results column headers
    for i, value in enumerate(col_headers):
        col_index = start_col + i
        absdiff_sheet.write(9, col_index, value, subsubtitle_merge_format)

    # out of bounds
    absdiff_sheet.set_column(0, 0, 4, oob_format)
    absdiff_sheet.set_column(absdiff_width + 3, 1000, 4, oob_format)
    for i in range(9):
        absdiff_sheet.set_row(i, None, oob_format)
    for i in range(10 + absdiff_len, 1000):
        absdiff_sheet.set_row(i, None, oob_format)

    # Column widths for values
    absdiff_sheet.set_column(2, 2, 15)
    absdiff_sheet.set_column(1, absdiff_width + 3, 12)

    # Row Heights
    absdiff_sheet.set_row(0, 25, oob_format)
    absdiff_sheet.set_row(1, 25, oob_format)
    absdiff_sheet.set_row(7, 31, oob_format)
    absdiff_sheet.set_row(8, 40, oob_format)
    absdiff_sheet.set_row(9, 40, oob_format)

    # Conditional Formatting
    absdiff_sheet.conditional_format(
        10,
        4,
        9 + absdiff_len,
        absdiff_width + 3,
        {
            "type": "text",
            "criteria": "containing",
            "value": "Identical",
            "format": identical_format,
        },
    )
    absdiff_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": ">=",
            "value": 0,
            "format": scientific_format,
        },
    )
    absdiff_sheet.conditional_format(
        10,
        4,
        9 + comp_len,
        comp_width + 3,
        {
            "type": "cell",
            "criteria": "<",
            "value": 0,
            "format": scientific_format,
        },
    )
    wb.close()

    # Copy the single post processed results into the comparison workbook
    target_file = outpath  # Target workbook filename
    target_sheet_names = [
        "Values",
        "Errors",
        "Statistical Checks",
    ]  # Target sheet names

    # Loop over the library source files
    for source_file in single_pp_files:
        copy_sheets_to_target(source_file, target_file, target_sheet_names)


def copy_sheets_to_target(source_filename, target_filename, target_sheet_names):
    """Performs copying of sheets between different workbooks

    Parameters
    ----------
    source_filename : str
        file for sheets to be copied from
    target_filename : str
        file for sheets to be copied to
    target_sheet_names : str
        name of sheets in target file
    """
    wb_target = openpyxl.load_workbook(target_filename)
    source_wb = openpyxl.load_workbook(source_filename)

    for source_sheet_name, target_sheet_name in zip(
        source_wb.sheetnames, target_sheet_names
    ):
        source_sheet = source_wb[source_sheet_name]
        target_sheet = wb_target.create_sheet(target_sheet_name)
        copy_sheet(source_sheet, target_sheet)

    wb_target.save(target_filename)


def copy_sheet(source_sheet: Worksheet, target_sheet: Worksheet) -> None:
    """Copy a sheet with style, format, layout, etc. from one Excel file
    to another Excel file.

    Parameters
    ----------
    source_sheet : Worksheet
        sheet to be copied
    target_sheet : Worksheet
        destination sheet
    """
    _copy_cells(source_sheet, target_sheet)  # copy all the cell values and styles
    _copy_sheet_attributes(source_sheet, target_sheet)


def _copy_sheet_attributes(source_sheet, target_sheet):
    """Copy contents of workbook sheets

    Parameters
    ----------
    source_sheet : str
        sheet name in source workbook
    target_sheet : str
        sheet name in target workbook
    """
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        print("Unable to copy default column wide")
    else:
        target_sheet.sheet_format.defaultColWidth = copy(
            source_sheet.sheet_format.defaultColWidth
        )

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(
            source_sheet.column_dimensions[key].min
        )  # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(
            source_sheet.column_dimensions[key].max
        )  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(
            source_sheet.column_dimensions[key].width
        )  # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(
            source_sheet.column_dimensions[key].hidden
        )


def _copy_cells(source_sheet, target_sheet):
    """Copy contents of cells including format in a workbook

    Parameters
    ----------
    source_sheet : str
        sheet name in source workbook
    target_sheet : str
        sheet name in target workbook
    """
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)
